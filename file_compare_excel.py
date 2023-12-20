import PySimpleGUI as sg
from openpyxl import load_workbook

def compare_files(file_path_a, file_path_b):
    '''
    Opens both files and checks if they are identical
    '''
    workbookA = load_workbook(filename=file_path_a)
    workbookB = load_workbook(filename=file_path_b)

    for sheetA, sheetB in zip(workbookA.worksheets, workbookB.worksheets):
        for rowA, rowB in zip(sheetA.iter_rows(), sheetB.iter_rows()):
            for cellA, cellB in zip(rowA, rowB):
                # Treat a cell with a value and a blank cell as different
                if (cellA.value is None and cellB.value is not None) or (cellA.value is not None and cellB.value is None):
                    sg.popup("Files are not exactly the same!")
                    return
                # Also check if the cell values are different
                elif cellA.value != cellB.value:
                    sg.popup("Files are not exactly the same!")
                    return
    sg.popup("Files are exactly the same!")



# define the layout of the GUI
layout = [
    [sg.Text("Select First File:")],
    [sg.Input('', key="fileA"), sg.FileBrowse(file_types=(("Excel Files", "*.xlsx"),))],
    [sg.Text("Select Second File: ")],
    [sg.Input('', key="fileB"), sg.FileBrowse(file_types=(("Excel Files", "*.xlsx"),))],
    [sg.Button("Check"), sg.Button("Exit")]
]

# create the window
window = sg.Window("File Comparison Checker", layout)

while True:
    # display and interact with window
    event, values = window.read()

    # if user closes window or clicks exit
    if event == sg.WINDOW_CLOSED or event == 'Exit':
        break

    # perform check
    if event == "Check":
        excel_file_A = values['fileA']
        excel_file_B = values['fileB']
        compare_files(excel_file_A, excel_file_B)


# close window
window.close()